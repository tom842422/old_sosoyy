import pywinauto

import time
from openpyxl import load_workbook

from pywinauto.keyboard import send_keys
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from time import sleep


def gbk2utf(in_data, tag):
    if tag == 1:
        return in_data.encode('gbk').decode('gbk')
    elif tag == 0:
        return in_data.encode('gbk').decode('gbk').encode('utf8')


options = webdriver.ChromeOptions()
prefs = {'download.prompt_for_download': True}
options.add_experimental_option('prefs', prefs)
browser = webdriver.Chrome(chrome_options=options)
browser.maximize_window()
# 全局等待
browser.implicitly_wait(20)
browser.get('https://srm.sosoyy.com/#/login')
# 账号
browser.find_element(By.XPATH, '//*[@id="form_item_mobile"]').click()
browser.find_element(By.XPATH, '//*[@id="form_item_mobile"]').send_keys(13518101893)
# 密码
browser.find_element(By.XPATH, '//*[@id="form_item_password"]').click()
browser.find_element(By.XPATH, '//*[@id="form_item_password"]').send_keys(123456)
# 验证码
browser.find_element(By.XPATH, '//*[@id="form_item_captcha"]').click()
browser.find_element(By.XPATH, '//*[@id="form_item_captcha"]').send_keys('2usw')
# 登录
browser.find_element(By.XPATH, '//*[@id="loginPage"]/div[2]/div/div[2]/button').click()
# 等待
sleep(2)

# 采购协同
browser.find_element(By.XPATH, '//*[@id="leftMenu"]/ul/li[2]/div/span/span/span').click()
sleep(1)
browser.find_element(By.CSS_SELECTOR, r'#sub_menu_1_\$\$_1-popup > li:nth-child(1) > span > a').click()

browser.implicitly_wait(5)
# 跳转上传采购计划
browser.find_element(By.XPATH, '/html/body/div[1]/div/div/div[1]/div[2]/ul/li[2]/ul/li[1]/span/a').click()
WebDriverWait(browser, 5, 1).until(EC.presence_of_element_located((By.ID, 'pageContent')))
# 重新定位当前页面
serach_windows = browser.current_window_handle
browser.implicitly_wait(5)
# 点击导入采购计划
browser.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div[2]/div[1]/div[4]/div[1]/div/button[1]').click()
sleep(3)
# #选择分组
browser.find_element(By.XPATH, '/html/body/div[12]/div/div[2]/di'
                               'v/div[2]/div[2]/div/div/div/div[1]/div[2]/div[3]/button[1]').click()
sleep(3)
# # 上传
browser.find_element(By.XPATH, '/html/body/div[12]/div/div[2]/div/div[2]/div[3]/button[2]').click()
# 上传文件
updata_file = pywinauto.Desktop()['打开']
# 选中地址输入框
updata_file['Toolbar3'].click()
# 输入路径
send_keys('D:\work_download')
send_keys("{ENTER}")
# 选中文件名输入框，输入文件名
updata_file["文件名(&N):Edit"].type_keys("采购计划-单正式.xlsx")
sleep(5)
# 上传
updata_file["打开(&O)"].click()
sleep(5)
# 确定
browser.find_element(By.XPATH, '/html/body/div[13]/div/div[2]/div/div[2]/div[3]/button[2]/span').click()
# 新建标签
new_windows = 'window.open("http://dddjds4067.sosoyy.com/Account/Index")'
browser.execute_script(new_windows)
handles = browser.window_handles
browser.switch_to.window(handles[1])
browser.implicitly_wait(20)
browser.find_element(By.XPATH, '//*[@id="mobile"]').click()
browser.find_element(By.XPATH, '//*[@id="mobile"]').send_keys(13518101801)
browser.find_element(By.XPATH, '//*[@id="password"]').click()
browser.find_element(By.XPATH, '//*[@id="password"]').send_keys(123456)
browser.find_element(By.XPATH, '//*[@id="loginForm"]/div[4]/div[1]/button').click()
sleep(200)
browser.refresh()
try:
    # 关闭买家给卖家的通知(后续添加多次关闭)
    browser.find_element(By.XPATH, '//*[@id="bjbBlock"]/div/div[1]/button/i').click()
    # 关闭微信弹窗
    browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div[1]/button/i').click()
except:
    # 关闭微信弹窗
    browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div[1]/button/i').click()
# 采购计划列表
browser.find_element(By.XPATH, '/html/body/div[1]/aside/section/ul/li[2]/ul/li[1]/a').click()
# 下载最新的计划
browser.find_element(By.XPATH,
                     '/html/body/div[1]/div[1]/section[2]/div/div/div[2]/div[1]/table/tbody/tr[1]/td[7]/a[2]').click()
sleep(1)
# 确认下载
browser.find_element(By.CSS_SELECTOR, '#downloadExecl').click()
# 选择保存路径和文件名
downfile = pywinauto.Desktop()['另存为']
sleep(1)
# 选中地址输入框
downfile["Toolbar3"].click()
sleep(2)
# 输入路径
down_filepath = 'D:\\work_download'
sleep(2)
send_keys(down_filepath)
send_keys("{ENTER}")
sleep(2)
# 获取当前时间
gettime = time.strftime('%Y%m%d%H%M', time.localtime(time.time()))
download_filename = f'计划-订单对接调试-瞻望-{gettime}.xlsx'
file_name = down_filepath + '\\' + download_filename
# 选中文件名输入框，输入文件名
send_keys(download_filename)
# downfile["文件名(&N):Edit"].type_keys(file_name)
# 保存
sleep(2)
downfile["保存(&S)"].click()
# 读取表头并找到价格字段写入保存
print(file_name)
sleep(5)
wb = load_workbook(file_name)
sheet = wb["Sheet1"]
max_column = sheet.max_column
max_row = sheet.max_row
for i in range(1, max_column + 1):
    t = sheet.cell(row=1, column=i).value
    if t == '价格':
        break
price = 11
for j in range(2, max_row + 1):
    cell_price = sheet.cell(row=j, column=i, value=price)
    price += 1
wb.save(filename=file_name)
sleep(5)
browser.find_element(By.CSS_SELECTOR,
                     'body > div.wrapper > div.content-wrapper > section.content > div > div > div:nth-child(2) > div.box-body > table > tbody > tr:nth-child(1) > td:nth-child(7) > a:nth-child(3)').click()
sleep(2)

# 选择文件
js = "document.getElementById('file').click()"
browser.execute_script(js)
sleep(3)
# 上传文件
upload = pywinauto.Desktop()['打开']
# 选中地址输入框
upload['Toolbar3'].click()
# 输入路径
send_keys(down_filepath)
send_keys("{ENTER}")
# 选中文件名输入框，输入文件名
upload["文件名&N:Edit"].click()
send_keys(download_filename)
sleep(4)
# 上传
upload["打开(&O)"].click()
# upload["打开(&O)"].click()
sleep(3)
# 下一步
browser.find_element(By.XPATH, '/html/body/div[1]/div[2]/div/div/div[3]/button[2]').click()
sleep(5)
# 切换到第一标签页
browser.switch_to.window(handles[0])
# 重置
browser.find_element(By.XPATH, '//*[@id="app"]/div/div[1]/div[2]/div[1]/div[2]/button[2]').click()
# 重置确认
browser.find_element(By.XPATH, '//*[@id="app"]/div/div[14]/div/p[2]/button[2]').click()
sleep(5)
# 提交订单
browser.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div/button[2]').click()
# 确认提交
browser.find_element(By.XPATH, '//*[@id="purchase-order-confirm"]/div[1]/div/div[3]/div/div[2]/div[3]').click()
# 确认弹窗
browser.find_element(By.XPATH, '/html/body/div[30]/div/div[2]/div/div[2]/div/div/div[2]/button').click()
# 跳转详情页
browser.find_element(By.XPATH,
                     '//*[@id="orderTable"]/div/div/div/div/div[1]/div/div[2]/table/tbody/tr[2]/td[12]/div/span[1]').click()
# 开票并保存确认
browser.find_element(By.XPATH,
                     '//*[@id="purchase-order-detail"]/div[1]/div/div[3]/div/div[1]/div[2]/div/div/button[1]').click()
browser.find_element(By.XPATH, '//*[@id="purchase-order-detail"]/div[1]/div/div[3]/div/div[3]/div[2]/button[2]').click()
browser.find_element(By.XPATH, '/html/body/div[29]/div/div[2]/div/div[2]/div/div/div[2]/button[2]').click()
